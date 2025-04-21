
import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment

def process_excel_file(input_path, reference_file, output_path):
    try:
        # Tenter de lire avec openpyxl (XLSX)
        try:
            df = pd.read_excel(input_path, engine='openpyxl')
        except Exception:
            # Si échec, essayer avec xlrd (XLS)
            df = pd.read_excel(input_path, engine='xlrd')
    except Exception as e:
        raise ValueError("Impossible de lire le fichier Excel. Assurez-vous qu'il est au bon format (.xls ou .xlsx).")

    # Nettoyage et traitement
    df.columns = [col.strip() for col in df.columns]
    df = df.rename(columns={
        "Nom complet": "Nom",
        "Téléphone": "Numero"
    })

    if "Numero" in df.columns:
        df["Numero"] = df["Numero"].astype(str).str.replace(r"[^0-9]", "", regex=True)

    # Chargement du modèle de référence
    ref_wb = openpyxl.load_workbook(reference_file)
    ref_ws = ref_wb.active

    wb = openpyxl.Workbook()
    ws = wb.active

    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    for col in ref_ws.column_dimensions:
        if col in ws.column_dimensions:
            ws.column_dimensions[col].width = ref_ws.column_dimensions[col].width

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="center")

    wb.save(output_path)
